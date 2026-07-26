"""
Turn an uploaded context document into plain text for the LLM.

Supported: PDF, CSV, TXT/MD, JSON, XLSX. Dispatch is by extension with a
sniff-based fallback, so a mislabelled upload still lands on the right reader.
"""

from __future__ import annotations

import csv
import io
import json
from pathlib import Path

import pdfplumber

MAX_CHARS = 120_000
"""Trim point for extracted text. Comfortably inside the model's window while
leaving room for the prompt and the JSON response."""


class UnsupportedDocument(Exception):
    pass


def _from_pdf(data: bytes) -> str:
    """
    Page text plus any detected tables.

    Financial disclosures put the numbers that matter inside ruled tables, and
    plain `extract_text` flattens those into ambiguous number soup. Appending a
    pipe-delimited rendering of each table preserves the row/column association
    the model needs to read a figure off the right line.
    """
    out: list[str] = []
    with pdfplumber.open(io.BytesIO(data)) as pdf:
        for i, page in enumerate(pdf.pages, start=1):
            out.append(f"\n--- page {i} ---")
            text = page.extract_text() or ""
            if text.strip():
                out.append(text)
            for table in page.extract_tables() or []:
                rows = [
                    " | ".join((c or "").replace("\n", " ").strip() for c in row)
                    for row in table
                    if any(c for c in row)
                ]
                if len(rows) > 1:
                    out.append("[table]\n" + "\n".join(rows))
    return "\n".join(out)


def _from_csv(data: bytes) -> str:
    text = data.decode("utf-8", errors="replace")
    try:
        dialect = csv.Sniffer().sniff(text[:4096])
    except csv.Error:
        dialect = csv.excel
    reader = csv.reader(io.StringIO(text), dialect)
    return "\n".join(" | ".join(cell.strip() for cell in row) for row in reader if row)


def _from_json(data: bytes) -> str:
    return json.dumps(json.loads(data.decode("utf-8", errors="replace")), indent=2)


def _from_xlsx(data: bytes) -> str:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # optional dependency
        raise UnsupportedDocument(
            "XLSX support needs openpyxl -- pip install openpyxl"
        ) from exc

    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    out: list[str] = []
    for ws in wb.worksheets:
        out.append(f"\n--- sheet: {ws.title} ---")
        for row in ws.iter_rows(values_only=True):
            if any(c is not None for c in row):
                out.append(" | ".join("" if c is None else str(c) for c in row))
    return "\n".join(out)


def _from_text(data: bytes) -> str:
    return data.decode("utf-8", errors="replace")


READERS = {
    ".pdf": _from_pdf,
    ".csv": _from_csv,
    ".tsv": _from_csv,
    ".json": _from_json,
    ".xlsx": _from_xlsx,
    ".xlsm": _from_xlsx,
    ".txt": _from_text,
    ".md": _from_text,
    ".text": _from_text,
}

SUPPORTED = sorted(READERS)


def extract_text(filename: str, data: bytes) -> str:
    """
    Extract plain text from an uploaded document.

    Raises UnsupportedDocument for unknown types, and for files that read as
    empty -- a scanned PDF with no text layer is better reported than silently
    turned into a blank report.
    """
    suffix = Path(filename).suffix.lower()
    reader = READERS.get(suffix)

    if reader is None:
        # Unknown extension: sniff the magic bytes before giving up.
        if data[:5] == b"%PDF-":
            reader = _from_pdf
        elif data[:2] == b"PK":
            reader = _from_xlsx
        else:
            try:
                data.decode("utf-8")
            except UnicodeDecodeError:
                raise UnsupportedDocument(
                    f"Unsupported file type '{suffix or filename}'. "
                    f"Supported: {', '.join(SUPPORTED)}"
                ) from None
            reader = _from_text

    text = reader(data).strip()

    if not text:
        raise UnsupportedDocument(
            f"No text could be extracted from '{filename}'. "
            "If this is a scanned PDF it has no text layer and would need OCR."
        )

    if len(text) > MAX_CHARS:
        head = text[: int(MAX_CHARS * 0.7)]
        tail = text[-int(MAX_CHARS * 0.3) :]
        text = f"{head}\n\n[... {len(text) - MAX_CHARS:,} characters trimmed ...]\n\n{tail}"

    return text
